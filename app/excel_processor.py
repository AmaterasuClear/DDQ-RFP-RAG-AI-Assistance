"""Process due diligence questionnaires in Excel format.

Reads questions from an Excel file, runs each through the RAG pipeline,
and writes formal answers back into a new column.
"""

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

from app.main import run_question
from app.storage import list_chunks


def _detect_question_column(headers: list[str]) -> int | None:
    """Find the most likely question column by scoring header names.

    Uses a weighted scoring system: longer headers containing question-related
    keywords score higher. Headers matching ID/index patterns are penalised.
    """
    id_patterns = [" id", " id ", " no ", " no.", "#", "ref", "code", "index"]
    question_keywords = [
        ("question text", 5),
        ("due diligence", 5),
        ("request item", 4),
        ("description", 3),
        ("question", 3),
        ("inquiry", 3),
        ("assessment", 2),
        ("criteria", 2),
        ("requirement", 2),
        ("ddq", 4),
        ("topic", 2),
        ("subject", 2),
        ("item", 1),
        ("query", 2),
        ("field", 1),
        ("section", 1),
        ("要求", 3),
        ("问题", 3),
        ("質問", 3),
    ]

    best_idx: int | None = None
    best_score = 0

    for idx, header in enumerate(headers):
        header_lower = header.strip().lower()
        if not header_lower:
            continue

        # Penalise ID/index columns
        is_id_col = any(pat in header_lower for pat in id_patterns)
        if is_id_col:
            continue

        score = 0
        for keyword, weight in question_keywords:
            if keyword in header_lower:
                score = max(score, weight)

        # Bonus for longer headers (more descriptive = more likely question column)
        if len(header_lower) > 15:
            score += 1

        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx


def _extract_questions(
    worksheet,
    question_col: int,
    header_row: int = 0,
    max_questions: int = 200,
) -> list[tuple[int, str]]:
    """Extract non-empty questions from the worksheet column.

    Returns list of (row, question_text) where row is 1-based openpyxl row number.
    """
    questions = []
    # header_row is 0-based; data starts at header_row + 2 in 1-based openpyxl rows
    for row in range(header_row + 2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=question_col + 1)
        text = str(cell.value or "").strip()
        if text and len(text) > 5:
            questions.append((row, text))
        if len(questions) >= max_questions:
            break
    return questions


def _read_worksheet(file_data: bytes, filename: str):
    """Read an Excel file and return a worksheet and workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel processing. Install with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(io.BytesIO(file_data), data_only=True)
    # Use the first visible worksheet
    sheet_names = [sn for sn in workbook.sheetnames if workbook[sn].sheet_state == "visible"]
    if not sheet_names:
        sheet_names = workbook.sheetnames
    worksheet = workbook[sheet_names[0]]
    return workbook, worksheet


def _get_header_row(worksheet, max_scan: int = 5) -> int:
    """Find the header row by looking for non-empty rows at the top."""
    for row_idx in range(min(max_scan, worksheet.max_row or 0)):
        row_values = [
            str(worksheet.cell(row=row_idx + 1, column=c + 1).value or "").strip()
            for c in range(min(worksheet.max_column or 0, 50))
        ]
        if any(v for v in row_values):
            return row_idx
    return 0


def process_questionnaire(
    file_data: bytes,
    filename: str,
    db_path: str | Path | None = None,
    progress_callback: Any = None,
    answer_column_name: str = "DDQ RAG Answer",
    confidence_column_name: str = "Confidence",
    evidence_column_name: str = "Source Evidence",
) -> tuple[bytes, str, dict[str, Any]]:
    """Process an Excel questionnaire.

    Args:
        file_data: Raw bytes of the Excel file.
        filename: Original filename.
        db_path: SQLite database path for RAG.
        progress_callback: Optional callable(current, total) for progress.
        answer_column_name: Header name for the answer column.
        confidence_column_name: Header name for the confidence column.
        evidence_column_name: Header name for the source evidence column.

    Returns:
        Tuple of (output_bytes, output_filename, summary_dict).
    """
    workbook, worksheet = _read_worksheet(file_data, filename)

    header_row = _get_header_row(worksheet)
    headers = [
        str(worksheet.cell(row=header_row + 1, column=c + 1).value or "").strip()
        for c in range(worksheet.max_column or 0)
    ]

    question_col = _detect_question_column(headers)
    if question_col is None:
        # Fall back to first column
        question_col = 0

    questions = _extract_questions(worksheet, question_col, header_row)

    if not questions:
        raise ValueError("No questions found in the worksheet. Ensure questions are in a single column.")

    # Find insertion columns (after the last used column)
    next_col = worksheet.max_column or len(headers)

    # Write answer, confidence, and evidence headers
    answer_col = next_col
    confidence_col = next_col + 1
    evidence_col = next_col + 2
    worksheet.cell(row=header_row + 1, column=answer_col + 1, value=answer_column_name)
    worksheet.cell(row=header_row + 1, column=confidence_col + 1, value=confidence_column_name)
    worksheet.cell(row=header_row + 1, column=evidence_col + 1, value=evidence_column_name)

    # Update header list for output
    headers.append(answer_column_name)
    headers.append(confidence_column_name)
    headers.append(evidence_column_name)

    total = len(questions)
    results_summary = {
        "total": total,
        "answered": 0,
        "high_confidence": 0,
        "medium_confidence": 0,
        "low_confidence": 0,
        "errors": 0,
        "questions": [],
    }

    for idx, (row_idx, question) in enumerate(questions):
        if progress_callback:
            progress_callback(idx + 1, total)

        try:
            result = run_question(question, db_path=db_path)

            # Write answer
            worksheet.cell(row=row_idx, column=answer_col + 1, value=result.response.answer)

            # Write confidence
            worksheet.cell(
                row=row_idx, column=confidence_col + 1,
                value=result.response.confidence_level,
            )

            # Write evidence — doc names, pages, and excerpts
            evidence_parts = []
            for cit in result.response.source_citations:
                evidence_parts.append(
                    f"[{cit.doc_name}, p.{cit.page}] {cit.quote[:200]}"
                )
            evidence_text = "\n".join(evidence_parts) if evidence_parts else (
                result.response.uncertainty_note or "No source evidence available."
            )
            worksheet.cell(row=row_idx, column=evidence_col + 1, value=evidence_text)

            results_summary["answered"] += 1
            conf = result.response.confidence_level
            if conf == "HIGH":
                results_summary["high_confidence"] += 1
            elif conf == "MEDIUM":
                results_summary["medium_confidence"] += 1
            else:
                results_summary["low_confidence"] += 1

            results_summary["questions"].append({
                "row": row_idx,
                "question": question[:120],
                "confidence": conf,
                "has_source": result.response.has_source,
            })

        except Exception as exc:
            results_summary["errors"] += 1
            worksheet.cell(
                row=row_idx, column=answer_col + 1,
                value=f"Error: {str(exc)[:200]}",
            )
            worksheet.cell(row=row_idx, column=confidence_col + 1, value="ERROR")
            worksheet.cell(row=row_idx, column=evidence_col + 1, value="")
            results_summary["questions"].append({
                "row": row_idx,
                "question": question[:120],
                "confidence": "ERROR",
                "error": str(exc)[:200],
            })

    # Adjust column widths for readability
    worksheet.column_dimensions[
        worksheet.cell(row=1, column=answer_col + 1).column_letter
    ].width = 50
    worksheet.column_dimensions[
        worksheet.cell(row=1, column=confidence_col + 1).column_letter
    ].width = 14
    worksheet.column_dimensions[
        worksheet.cell(row=1, column=evidence_col + 1).column_letter
    ].width = 40

    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    output_filename = Path(filename).stem + "_answered.xlsx"

    return output.getvalue(), output_filename, results_summary


def read_questions_only(
    file_data: bytes,
    filename: str,
) -> list[str]:
    """Extract questions from an Excel file without processing."""
    workbook, worksheet = _read_worksheet(file_data, filename)
    header_row = _get_header_row(worksheet)
    headers = [
        str(worksheet.cell(row=header_row + 1, column=c + 1).value or "").strip()
        for c in range(worksheet.max_column or 0)
    ]
    question_col = _detect_question_column(headers)
    if question_col is None:
        question_col = 0

    questions = _extract_questions(worksheet, question_col, header_row)
    return [q for _, q in questions]