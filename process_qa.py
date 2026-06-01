"""Quick batch processor for Excel DDQ."""
import sys, os, io, time
from pathlib import Path

os.environ.setdefault("HF_HUB_OFFLINE", "1")
os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")

env_path = Path(__file__).resolve().parent / ".env"
if env_path.exists():
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            k, v = k.strip(), v.strip().strip('"').strip("'")
            if k and k not in os.environ:
                os.environ[k] = v

os.environ["DDQ_GENERATION_PROVIDER"] = "gemini"

try:
    import pyarrow
except ImportError:
    pass

from app.embeddings import generate_embedding
from app.storage import init_db, list_chunks
from app.config import DEFAULT_DB_PATH
from app.main import run_question

_warm = generate_embedding("Initialize model before batch processing.")
init_db(DEFAULT_DB_PATH)
_chunks = list_chunks(DEFAULT_DB_PATH)
print(f"Model loaded, {len(_chunks)} chunks in DB. Processing...")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("test-DDQ.xlsx")
ws = wb.active

headers = [str(ws.cell(row=1, column=c + 1).value or "").strip()
           for c in range(ws.max_column or 0)]
# Use the smart detection from excel_processor (finds question TEXT column, not ID column)
from app.excel_processor import _detect_question_column
question_col = _detect_question_column(headers)
if question_col is None:
    question_col = 1

next_col = ws.max_column
ac, cc, ec = next_col, next_col + 1, next_col + 2
ws.cell(row=1, column=ac + 1, value="DDQ RAG Answer")
ws.cell(row=1, column=cc + 1, value="Confidence")
ws.cell(row=1, column=ec + 1, value="Source Evidence")

total = answered = errors = high = medium = low = 0

for row in range(2, ws.max_row + 1):
    question = str(ws.cell(row=row, column=question_col + 1).value or "").strip()
    if not question or len(question) < 5:
        continue
    total += 1
    qid = str(ws.cell(row=row, column=1).value or "").strip()
    print(f"[{total}/37] {qid}...", end=" ", flush=True)
    try:
        result = run_question(question)
        ws.cell(row=row, column=ac + 1, value=result.response.answer)
        ws.cell(row=row, column=cc + 1, value=result.response.confidence_level)

        evidence_parts = []
        for cit in result.response.source_citations:
            evidence_parts.append(f"[{cit.doc_name}, p.{cit.page}] {cit.quote[:200]}")
        evidence_text = "\n".join(evidence_parts) if evidence_parts else (
            result.response.uncertainty_note or "No source evidence.")
        ws.cell(row=row, column=ec + 1, value=evidence_text)

        conf = result.response.confidence_level
        if conf == "HIGH":
            high += 1
        elif conf == "MEDIUM":
            medium += 1
        else:
            low += 1
        answered += 1
        print(conf)
    except Exception as e:
        ws.cell(row=row, column=ac + 1, value=f"Error: {str(e)[:200]}")
        ws.cell(row=row, column=cc + 1, value="ERROR")
        errors += 1
        print(f"ERROR: {e}")

    time.sleep(1.0)

ws.column_dimensions[get_column_letter(ac + 1)].width = 50
ws.column_dimensions[get_column_letter(cc + 1)].width = 14
ws.column_dimensions[get_column_letter(ec + 1)].width = 40

wb.save("test-DDQ_answered.xlsx")
print(f"\nDone! {answered} answered, {errors} errors | HIGH={high} MEDIUM={medium} LOW={low}")
